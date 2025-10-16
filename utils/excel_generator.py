import tempfile
import os
import json
from pathlib import Path
import re

def load_course_categories():
    """FUTURE-PROOF VERSION: Load course categories from separate JSON files."""
    course_data_dir = Path(__file__).parent.parent / "course_data"
    
    categories = {
        "ie_core": {},
        "technical_electives": {},
        "gen_ed": {
            "wellness": {},
            "entrepreneurship": {},
            "language_communication": {},
            "thai_citizen_global": {},
            "aesthetics": {}
        },
        "all_courses": {}
    }
    
    # FUTURE-PROOF: Find all B-IE files dynamically
    ie_files = []
    if course_data_dir.exists():
        for json_file in course_data_dir.glob("B-IE-*.json"):
            year_match = re.search(r'B-IE-(\d{4})\.json', json_file.name)
            if year_match:
                year = int(year_match.group(1))
                ie_files.append((year, json_file))
    
    # Sort by year (newest first) and process
    ie_files.sort(key=lambda x: x[0], reverse=True)
    
    # Load IE Core courses from available B-IE files
    for year, ie_file in ie_files:
        try:
            with open(ie_file, 'r', encoding='utf-8') as f:
                ie_data = json.load(f)
                
                # Process industrial_engineering_courses
                for course in ie_data.get("industrial_engineering_courses", []):
                    if course["code"] not in categories["all_courses"]:
                        if course.get("technical_electives", False):
                            categories["technical_electives"][course["code"]] = course
                        else:
                            categories["ie_core"][course["code"]] = course
                        categories["all_courses"][course["code"]] = course
                
                # Process other_related_courses
                for course in ie_data.get("other_related_courses", []):
                    if course["code"] not in categories["all_courses"]:
                        categories["ie_core"][course["code"]] = course  
                        categories["all_courses"][course["code"]] = course
                        
        except Exception as e:
            print(f"Error loading {ie_file}: {e}")
            continue
    
    # Load Gen-Ed courses (unchanged)
    gen_ed_file = course_data_dir / "gen_ed_courses.json"
    if gen_ed_file.exists():
        try:
            with open(gen_ed_file, 'r', encoding='utf-8') as f:
                gen_ed_data = json.load(f)
                gen_ed_courses = gen_ed_data.get("gen_ed_courses", {})
                for subcategory in ["wellness", "entrepreneurship", "language_communication", "thai_citizen_global", "aesthetics"]:
                    for course in gen_ed_courses.get(subcategory, []):
                        categories["gen_ed"][subcategory][course["code"]] = course
                        categories["all_courses"][course["code"]] = course
        except Exception as e:
            print(f"Error loading gen_ed_courses.json: {e}")
    
    return categories

def classify_course(course_code, course_name="", course_categories=None):
    """
    Classify course into appropriate category using loaded JSON files.
    PRIORITY ORDER: Gen-Ed → Technical Electives → IE Core → Free Electives
    Returns: (category, subcategory, is_identified)
    """
    if course_categories is None:
        course_categories = load_course_categories()
    
    code = course_code.upper()
    
    # PRIORITY 1: Check Gen-Ed courses FIRST (highest priority)
    for subcategory, courses in course_categories["gen_ed"].items():
        if code in courses:
            return ("gen_ed", subcategory, True)
    
    # PRIORITY 2: Check Technical Electives
    if code in course_categories["technical_electives"]:
        return ("technical_electives", "technical", True)
    
    # PRIORITY 3: Check IE Core courses
    if code in course_categories["ie_core"]:
        return ("ie_core", "core", True)
    
    # PRIORITY 4: Everything else is free elective (not in our database)
    return ("free_electives", "free", False)  # False = not identified in database

def create_smart_registration_excel(student_info, semesters, validation_results):
    """
    Create a smart, dynamic Excel registration format with proper course detection.
    FIXED: Now properly handles technical electives from B-IE files.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registration Plan"
        
        # Load course categories once (now includes technical electives from B-IE files)
        course_categories = load_course_categories()
        
        # Define styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        small_font = Font(size=8)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Color fills
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # For unidentified
        purple_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  # For validation issues
        
        # Set column widths
        column_widths = {
            'A': 12, 'B': 40, 'C': 8, 'D': 8, 'E': 15, 'F': 40, 'G': 8, 'H': 8,
            'I': 15, 'J': 40, 'K': 8, 'L': 8, 'M': 15, 'N': 40, 'O': 8, 'P': 8
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # HEADER SECTION
        ws['A1'] = "KU INDUSTRIAL ENGINEERING - SMART COURSE REGISTRATION ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:P1')
        ws['A1'].alignment = center_align
        ws['A1'].fill = blue_fill
        
        # Student Info
        ws['A3'] = "Student ID:"
        ws['B3'] = student_info.get('id', '')
        ws['B3'].fill = yellow_fill
        ws['E3'] = "Name:"
        ws['F3'] = student_info.get('name', '')
        ws['F3'].fill = yellow_fill
        
        # CLASSIFY ALL COURSES WITH PROPER DETECTION
        classified_courses = {
            "ie_core": [],
            "gen_ed": {
                "wellness": [],
                "entrepreneurship": [],
                "language_communication": [],
                "thai_citizen_global": [],
                "aesthetics": []
            },
            "technical_electives": [],
            "free_electives": [],
            "unidentified": []  # Track unidentified courses
        }
        
        unidentified_count = 0
        validation_issues_count = 0
        
        # Create validation lookup for faster access
        validation_lookup = {}
        for result in validation_results:
            course_code = result.get('course_code')
            if course_code and course_code != 'CREDIT_LIMIT':
                validation_lookup[course_code] = {
                    'is_valid': result.get('is_valid', True),
                    'reason': result.get('reason', '')
                }
        
        # Process all courses from all semesters
        for sem_idx, semester in enumerate(semesters):
            semester_name = semester.get("semester", f"Semester {sem_idx + 1}")
            year = semester.get("year_int", 0)
            semester_type = semester.get("semester_type", "")
            
            for course in semester.get("courses", []):
                course_code = course.get("code", "")
                course_name = course.get("name", "")
                grade = course.get("grade", "")
                credits = course.get("credits", 0)
                
                # Check validation status
                validation_info = validation_lookup.get(course_code, {'is_valid': True, 'reason': ''})
                is_valid = validation_info['is_valid']
                issue_reason = validation_info['reason']
                
                if not is_valid:
                    validation_issues_count += 1
                
                # Classify course with FIXED detection (technical electives now properly detected)
                category, subcategory, is_identified = classify_course(course_code, course_name, course_categories)
                
                if not is_identified:
                    unidentified_count += 1
                
                course_info = {
                    "code": course_code,
                    "name": course_name,
                    "grade": grade,
                    "credits": credits,
                    "semester": semester_name,
                    "is_valid": is_valid,
                    "issue": issue_reason,
                    "year": year,
                    "semester_type": semester_type,
                    "is_identified": is_identified
                }
                
                # Place in appropriate category (FIXED: technical electives now properly classified)
                if category == "ie_core":
                    classified_courses["ie_core"].append(course_info)
                elif category == "gen_ed":
                    classified_courses["gen_ed"][subcategory].append(course_info)
                elif category == "technical_electives":
                    classified_courses["technical_electives"].append(course_info)
                elif category == "unidentified":
                    classified_courses["unidentified"].append(course_info)
                else:
                    classified_courses["free_electives"].append(course_info)
        
        # Add system status warnings
        current_row = 4
        if unidentified_count > 0 or validation_issues_count > 0:
            ws[f'A{current_row}'] = "⚠️ SYSTEM ALERTS:"
            ws[f'A{current_row}'].fill = orange_fill
            ws[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            if unidentified_count > 0:
                ws[f'A{current_row}'] = f"🔍 DATABASE EXPANSION NEEDED: {unidentified_count} new courses found"
                ws[f'A{current_row}'].fill = orange_fill
                ws[f'A{current_row}'].font = Font(bold=True)
                ws.merge_cells(f'A{current_row}:P{current_row}')
                current_row += 1
            
            if validation_issues_count > 0:
                ws[f'A{current_row}'] = f"❌ VALIDATION ISSUES: {validation_issues_count} courses have prerequisite problems"
                ws[f'A{current_row}'].fill = red_fill
                ws[f'A{current_row}'].font = Font(bold=True)
                ws.merge_cells(f'A{current_row}:P{current_row}')
                current_row += 1
        
        current_row += 1
        
        # Add technical electives info
        tech_electives_count = len(classified_courses["technical_electives"])
        if tech_electives_count > 0:
            ws[f'A{current_row}'] = f"✅ Enhanced: {tech_electives_count} courses properly classified as Technical Electives"
            ws[f'A{current_row}'].fill = green_fill
            ws[f'A{current_row}'].font = Font(bold=True)
            ws.merge_cells(f'A{current_row}:P{current_row}')
            current_row += 1
        
        current_row += 1
        
        # Function to add course rows with comprehensive status detection
        def add_category_section(title, courses, start_row, color_fill=None, required_credits=None):
            nonlocal current_row
            
            # Calculate earned credits
            earned_credits = sum(c["credits"] for c in courses if c["grade"] not in ['F', 'W', 'N', ''])
            
            # Section header with credit info
            header_text = title
            if required_credits:
                header_text += f" - Required: {required_credits} credits, Earned: {earned_credits}"
            else:
                header_text += f" - Earned: {earned_credits} credits"
            
            ws[f'A{start_row}'] = header_text
            ws[f'A{start_row}'].font = header_font
            if color_fill:
                ws[f'A{start_row}'].fill = color_fill
            else:
                ws[f'A{start_row}'].fill = blue_fill
            ws.merge_cells(f'A{start_row}:P{start_row}')
            current_row = start_row + 1
            
            if not courses:
                ws[f'A{current_row}'] = "No courses in this category"
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].fill = gray_fill
                ws[f'A{current_row}'].alignment = center_align
                ws.merge_cells(f'A{current_row}:P{current_row}')
                return current_row + 2
            
            # Column headers
            headers = ["Code", "Course Name", "Grade", "Credits", "Semester", "Status"]
            for j, header in enumerate(headers):
                if j < 6:  # Only show first 6 columns
                    cell = ws[f'{chr(65 + j)}{current_row}']
                    cell.value = header
                    cell.font = subheader_font
                    cell.alignment = center_align
                    cell.fill = gray_fill
                    cell.border = border
            current_row += 1
            
            # Add courses
            for course in courses:
                ws[f'A{current_row}'] = course["code"]
                ws[f'A{current_row}'].border = border
                ws[f'A{current_row}'].font = small_font
                
                ws[f'B{current_row}'] = course["name"]
                ws[f'B{current_row}'].border = border
                ws[f'B{current_row}'].font = small_font
                ws[f'B{current_row}'].alignment = left_align
                
                ws[f'C{current_row}'] = course["grade"]
                ws[f'C{current_row}'].border = border
                ws[f'C{current_row}'].font = small_font
                ws[f'C{current_row}'].alignment = center_align
                
                ws[f'D{current_row}'] = course["credits"]
                ws[f'D{current_row}'].border = border
                ws[f'D{current_row}'].font = small_font
                ws[f'D{current_row}'].alignment = center_align
                
                ws[f'E{current_row}'] = course["semester"]
                ws[f'E{current_row}'].border = border
                ws[f'E{current_row}'].font = small_font
                
                # Status column with comprehensive information
                status_text = ""
                row_color = None
                
                if not course["is_identified"]:
                    status_text = "NEW COURSE - NEEDS CLASSIFICATION"
                    row_color = orange_fill
                elif not course["is_valid"]:
                    status_text = f"INVALID: {course['issue']}"
                    row_color = red_fill
                elif course["grade"] in ["A", "B+", "B", "C+", "C", "D+", "D", "P"]:
                    status_text = "COMPLETED"
                    row_color = green_fill
                elif course["grade"] == "F":
                    status_text = "FAILED"
                    row_color = red_fill
                elif course["grade"] == "W":
                    status_text = "WITHDRAWN"
                    row_color = yellow_fill
                elif course["grade"] in ["N", ""]:
                    status_text = "IN PROGRESS"
                    row_color = yellow_fill
                else:
                    status_text = f"GRADE: {course['grade']}"
                
                ws[f'F{current_row}'] = status_text
                ws[f'F{current_row}'].border = border
                ws[f'F{current_row}'].font = small_font
                
                # Apply row coloring
                if row_color:
                    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                        ws[f'{col}{current_row}'].fill = row_color
                
                current_row += 1
            
            return current_row + 1
        
        # IE CORE COURSES SECTION
        current_row = add_category_section(
            "IE CORE COURSES", 
            classified_courses["ie_core"], 
            current_row,
            blue_fill,
            "110"
        )
        
        # TECHNICAL ELECTIVES SECTION (FIXED: Now properly populated from B-IE files)
        current_row = add_category_section(
            "TECHNICAL ELECTIVES (Enhanced: From B-IE Files)",
            classified_courses["technical_electives"],
            current_row,
            blue_fill
        )
        
        # UNIDENTIFIED COURSES SECTION (HIGH PRIORITY)
        if classified_courses["unidentified"]:
            current_row = add_category_section(
                "🔍 NEW COURSES - REQUIRE DATABASE EXPANSION",
                classified_courses["unidentified"],
                current_row,
                orange_fill
            )
        
        # GEN-ED SECTIONS with proper credit requirements
        gen_ed_categories = [
            ("wellness", "WELLNESS (Gen-Ed)", "7"),
            ("entrepreneurship", "ENTREPRENEURSHIP (Gen-Ed)", "3"),
            ("language_communication", "LANGUAGE & COMMUNICATION (Gen-Ed)", "15"),
            ("thai_citizen_global", "THAI CITIZEN & GLOBAL CITIZENSHIP (Gen-Ed)", "2"),
            ("aesthetics", "AESTHETICS (Gen-Ed)", "3")
        ]
        
        for category_key, category_name, required_credits in gen_ed_categories:
            courses = classified_courses["gen_ed"][category_key]
            current_row = add_category_section(
                category_name,
                courses,
                current_row,
                gray_fill,
                required_credits
            )
        
        # FREE ELECTIVES SECTION (should now have fewer courses since technical electives are properly classified)
        current_row = add_category_section(
            "FREE ELECTIVES (Enhanced: Excludes Technical Electives)",
            classified_courses["free_electives"],
            current_row,
            yellow_fill
        )
        
        # COMPREHENSIVE SUMMARY SECTION
        current_row += 1
        
        ws[f'A{current_row}'] = "COMPREHENSIVE CREDIT ANALYSIS"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = blue_fill
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1
        
        # Calculate credits by category with completion rates
        ie_credits = sum(c["credits"] for c in classified_courses["ie_core"] if c["grade"] not in ['F', 'W', 'N', ''])
        ie_total = sum(c["credits"] for c in classified_courses["ie_core"])
        
        gen_ed_credits = {}
        gen_ed_totals = {}
        
        for subcategory in ["wellness", "entrepreneurship", "language_communication", "thai_citizen_global", "aesthetics"]:
            gen_ed_credits[subcategory] = sum(
                c["credits"] for c in classified_courses["gen_ed"][subcategory] 
                if c["grade"] not in ['F', 'W', 'N', '']
            )
            gen_ed_totals[subcategory] = sum(
                c["credits"] for c in classified_courses["gen_ed"][subcategory]
            )
        
        # FIXED: Technical electives now properly separated from free electives
        tech_credits = sum(c["credits"] for c in classified_courses["technical_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
        free_credits = sum(c["credits"] for c in classified_courses["free_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
        unidentified_credits = sum(c["credits"] for c in classified_courses["unidentified"] if c["grade"] not in ['F', 'W', 'N', ''])
        
        # Credit requirements mapping
        requirements = {
            "IE Core": (110, ie_credits),
            "Wellness": (7, gen_ed_credits["wellness"]),
            "Entrepreneurship": (3, gen_ed_credits["entrepreneurship"]),
            "Language & Communication": (15, gen_ed_credits["language_communication"]),
            "Thai Citizen & Global": (2, gen_ed_credits["thai_citizen_global"]),
            "Aesthetics": (3, gen_ed_credits["aesthetics"]),
            "Technical Electives": (None, tech_credits),
            "Free Electives": (None, free_credits),
            "New Courses": (None, unidentified_credits)
        }
        
        # Summary headers
        headers = ["Category", "Required", "Earned", "Status", "Notes"]
        for j, header in enumerate(headers):
            cell = ws[f'{chr(65 + j)}{current_row}']
            cell.value = header
            cell.font = subheader_font
            cell.alignment = center_align
            cell.fill = gray_fill
            cell.border = border
        current_row += 1
        
        total_required = 0
        total_earned = 0
        
        for category, (required, earned) in requirements.items():
            ws[f'A{current_row}'] = category
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].border = border
            
            ws[f'B{current_row}'] = required if required else "Variable"
            ws[f'B{current_row}'].font = normal_font
            ws[f'B{current_row}'].border = border
            ws[f'B{current_row}'].alignment = center_align
            
            ws[f'C{current_row}'] = earned
            ws[f'C{current_row}'].font = normal_font
            ws[f'C{current_row}'].border = border
            ws[f'C{current_row}'].alignment = center_align
            
            # Status determination
            status = "N/A"
            status_color = None
            notes = ""
            
            if required:
                total_required += required
                total_earned += earned
                
                if earned >= required:
                    status = "✅ COMPLETE"
                    status_color = green_fill
                elif earned > 0:
                    status = "⚠️ PARTIAL"
                    status_color = yellow_fill
                    notes = f"Need {required - earned} more"
                else:
                    status = "❌ NOT STARTED"
                    status_color = red_fill
                    notes = f"Need {required} credits"
            else:
                if earned > 0:
                    status = f"✅ {earned} credits"
                    status_color = green_fill
                else:
                    status = "No courses"
            
            # Special handling for unidentified
            if category == "New Courses" and earned > 0:
                status = "⚠️ NEEDS CLASSIFICATION"
                status_color = orange_fill
                notes = f"{len(classified_courses['unidentified'])} courses need categorization"
            
            # Special note for enhanced categories
            if category in ["Technical Electives", "Free Electives"]:
                if category == "Technical Electives":
                    notes += f" - {len(classified_courses['technical_electives'])} courses from B-IE files"
                elif category == "Free Electives":
                    notes += " - Excludes technical electives"
            
            ws[f'D{current_row}'] = status
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].border = border
            if status_color:
                ws[f'D{current_row}'].fill = status_color
            
            ws[f'E{current_row}'] = notes
            ws[f'E{current_row}'].font = small_font
            ws[f'E{current_row}'].border = border
            
            current_row += 1
        
        # Grand total
        current_row += 1
        ws[f'A{current_row}'] = "TOTAL GRADUATION PROGRESS"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].fill = blue_fill
        
        total_with_unidentified = total_earned + unidentified_credits
        
        ws[f'B{current_row}'] = "140 (Est.)"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].fill = blue_fill
        
        ws[f'C{current_row}'] = f"{total_with_unidentified} ({total_earned} + {unidentified_credits} new)"
        ws[f'C{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'].fill = blue_fill
        
        completion_rate = (total_with_unidentified / 140) * 100
        ws[f'D{current_row}'] = f"{completion_rate:.1f}% Complete"
        ws[f'D{current_row}'].font = Font(bold=True)
        if completion_rate >= 100:
            ws[f'D{current_row}'].fill = green_fill
        elif completion_rate >= 80:
            ws[f'D{current_row}'].fill = yellow_fill
        else:
            ws[f'D{current_row}'].fill = red_fill
        
        # System recommendations
        current_row += 3
        ws[f'A{current_row}'] = "SYSTEM RECOMMENDATIONS"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = blue_fill
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1
        
        recommendations = []
        
        recommendations.append(f"✅ Enhanced: Technical electives now properly classified from B-IE files ({len(classified_courses['technical_electives'])} courses)")
        
        if unidentified_count > 0:
            recommendations.append(f"🔍 PRIORITY: Classify {unidentified_count} new courses to get accurate analysis")
        
        if validation_issues_count > 0:
            recommendations.append(f"❌ Fix {validation_issues_count} prerequisite issues")
        
        for category, (required, earned) in requirements.items():
            if required and earned < required:
                recommendations.append(f"📚 Complete {category}: need {required - earned} more credits")
        
        if len(recommendations) == 1:  # Only the FIXED message
            recommendations.append("✅ All requirements appear to be on track!")
        
        for i, rec in enumerate(recommendations):
            ws[f'A{current_row + i}'] = rec
            ws[f'A{current_row + i}'].font = normal_font
            if "Enhanced" in rec:
                ws[f'A{current_row + i}'].fill = green_fill
            elif "PRIORITY" in rec or "new courses" in rec:
                ws[f'A{current_row + i}'].fill = orange_fill
            elif "Fix" in rec:
                ws[f'A{current_row + i}'].fill = red_fill
            elif "Complete" in rec:
                ws[f'A{current_row + i}'].fill = yellow_fill
            else:
                ws[f'A{current_row + i}'].fill = green_fill
        
        # Save to bytes
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb.save(tmp_file.name)
            
            with open(tmp_file.name, 'rb') as f:
                excel_bytes = f.read()
            
            os.unlink(tmp_file.name)
            return excel_bytes, unidentified_count
            
    except Exception as e:
        raise Exception(f"Error creating Excel file: {e}")
